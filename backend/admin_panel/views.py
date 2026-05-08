"""admin_panel/views.py"""
import logging
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from core.permissions import IsAdmin
from users.models import User
from users.serializers import UserAdminSerializer
from products.models import Product
from orders.models import Order
from sellers.models import SellerProfile

logger = logging.getLogger(__name__)


class AdminUserListView(generics.ListAPIView):
    """Admin: list all users with filtering."""
    serializer_class = UserAdminSerializer
    permission_classes = [IsAdmin]
    search_fields = ['email', 'name', 'phone']
    filterset_fields = ['role', 'verified', 'is_active']

    def get_queryset(self):
        return User.objects.all().order_by('-date_joined')


class AdminUserDetailView(generics.RetrieveUpdateAPIView):
    """Admin: retrieve or update any user."""
    serializer_class = UserAdminSerializer
    permission_classes = [IsAdmin]
    queryset = User.objects.all()


class AdminSuspendUserView(APIView):
    """Admin: suspend (deactivate) a user account."""
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'status': 'error', 'message': 'User not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        if user.role == 'admin':
            return Response(
                {'status': 'error', 'message': 'Cannot suspend an admin account.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        user.is_active = False
        user.save(update_fields=['is_active'])
        logger.warning('Admin %s suspended user %s', request.user.email, user.email)
        return Response({
            'status': 'success',
            'message': f'User {user.email} has been suspended.',
        })


class AdminActivateUserView(APIView):
    """Admin: re-activate a suspended user account."""
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'status': 'error', 'message': 'User not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        user.is_active = True
        user.save(update_fields=['is_active'])
        return Response({
            'status': 'success',
            'message': f'User {user.email} has been activated.',
        })


class AdminVerifySellerView(APIView):
    """Admin: verify a seller profile."""
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            profile = SellerProfile.objects.get(pk=pk)
        except SellerProfile.DoesNotExist:
            return Response(
                {'status': 'error', 'message': 'Seller profile not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        profile.verified = True
        profile.save(update_fields=['verified'])
        return Response({
            'status': 'success',
            'message': f'Seller "{profile.business_name}" verified.',
        })


class AdminPlatformStatsView(APIView):
    """Admin: high-level platform statistics dashboard."""
    permission_classes = [IsAdmin]

    def get(self, request):
        now = timezone.now()
        thirty_days_ago = now - timedelta(days=30)

        total_users = User.objects.count()
        buyers = User.objects.filter(role='buyer').count()
        sellers = User.objects.filter(role='seller').count()
        new_users_30d = User.objects.filter(date_joined__gte=thirty_days_ago).count()

        total_products = Product.objects.count()
        active_products = Product.objects.filter(is_active=True).count()

        total_orders = Order.objects.count()
        revenue_data = Order.objects.filter(
            status__in=['processing', 'shipped', 'delivered']
        ).aggregate(total=Sum('total_amount'), avg_order=Avg('total_amount'))

        orders_30d = Order.objects.filter(created_at__gte=thirty_days_ago).count()

        return Response({
            'status': 'success',
            'data': {
                'users': {
                    'total': total_users,
                    'buyers': buyers,
                    'sellers': sellers,
                    'new_last_30_days': new_users_30d,
                },
                'products': {
                    'total': total_products,
                    'active': active_products,
                },
                'orders': {
                    'total': total_orders,
                    'last_30_days': orders_30d,
                    'total_revenue': str(revenue_data['total'] or 0),
                    'average_order_value': str(
                        round(revenue_data['avg_order'] or 0, 2)
                    ),
                },
            },
        })


class AdminOrderListView(generics.ListAPIView):
    """Admin: list all orders."""
    permission_classes = [IsAdmin]

    def get_queryset(self):
        return Order.objects.select_related('buyer').prefetch_related(
            'items__product'
        ).order_by('-created_at')

    def list(self, request, *args, **kwargs):
        from orders.serializers import OrderSerializer
        queryset = self.get_queryset()
        serializer = OrderSerializer(queryset, many=True)
        return Response({'status': 'success', 'count': queryset.count(), 'data': serializer.data})


class AdminAnalyticsView(APIView):
    """
    Admin: Full analytics dashboard.
    GET /api/admin/analytics
    """
    permission_classes = [IsAdmin]

    def get(self, request):
        from django.db.models.functions import TruncDate
        now = timezone.now()
        thirty_days_ago = now - timedelta(days=30)

        # Orders per day (last 30 days)
        orders_per_day = (
            Order.objects.filter(created_at__gte=thirty_days_ago)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'), revenue=Sum('total_amount'))
            .order_by('day')
        )

        # Top-selling products by quantity
        from orders.models import OrderItem
        top_products = (
            OrderItem.objects.values('product__id', 'product__name')
            .annotate(total_sold=Sum('quantity'))
            .order_by('-total_sold')[:10]
        )

        # Seller performance
        seller_performance = (
            SellerProfile.objects.select_related('user')
            .annotate(
                total_products=Count('user__products'),
                avg_rating=Avg('user__products__reviews__stars'),
            )
            .values(
                'business_name',
                'user__email',
                'total_products',
                'avg_rating',
                'rating_avg',
            )
            .order_by('-rating_avg')[:10]
        )

        # Revenue summary
        revenue = Order.objects.filter(
            status__in=['processing', 'shipped', 'delivered']
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id'),
            avg=Avg('total_amount'),
        )

        return Response({
            'status': 'success',
            'data': {
                'revenue': {
                    'total': str(revenue['total'] or 0),
                    'order_count': revenue['count'],
                    'avg_order_value': str(round(revenue['avg'] or 0, 2)),
                },
                'orders_per_day': list(orders_per_day),
                'top_selling_products': list(top_products),
                'top_sellers': list(seller_performance),
            },
        })


class AdminReportedContentView(generics.ListAPIView):
    """
    Admin: List all unresolved content reports.
    GET /api/admin/reported-content
    """
    permission_classes = [IsAdmin]

    def get_queryset(self):
        from reviews.models import ContentReport
        return ContentReport.objects.filter(resolved=False).select_related(
            'reporter', 'review', 'thread'
        ).order_by('-created_at')

    def list(self, request, *args, **kwargs):
        from reviews.serializers import ContentReportSerializer
        queryset = self.get_queryset()
        serializer = ContentReportSerializer(queryset, many=True)
        return Response({
            'status': 'success',
            'count': queryset.count(),
            'results': serializer.data,
        })


class AdminAssistedUploadView(APIView):
    """
    Admin: Assisted upload service (FR-02.10).
    Allows an admin to upload an .xlsx file on behalf of a specific seller.
    """
    permission_classes = [IsAdmin]
    # parser_classes = [MultiPartParser]

    def post(self, request):
        from products.models import Product, Category
        from products.serializers import ProductBulkRowSerializer
        import openpyxl

        file = request.FILES.get('file')
        seller_id = request.data.get('seller_id')

        if not file or not seller_id:
            return Response(
                {'status': 'error', 'message': 'file and seller_id are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
            
        try:
            seller = User.objects.get(id=seller_id, role='seller')
        except User.DoesNotExist:
            return Response(
                {'status': 'error', 'message': 'Seller not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                {'status': 'error', 'message': 'Only .xlsx files are accepted.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response(
                {'status': 'error', 'message': f'Could not parse file: {e}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = ['name', 'description', 'price', 'stock_qty', 'SKU', 'category']
        missing = [h for h in expected if h not in headers]
        if missing:
            return Response(
                {'status': 'error', 'message': f'Missing columns: {missing}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        col_map = {v: i for i, v in enumerate(headers)}
        created, errors = [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_data = {
                'name': row[col_map['name']],
                'description': row[col_map.get('description', -1)] or '',
                'price': row[col_map['price']],
                'stock_qty': row[col_map['stock_qty']],
                'SKU': row[col_map['SKU']],
                'category': row[col_map.get('category', -1)] or '',
            }
            serializer = ProductBulkRowSerializer(data=row_data)
            if serializer.is_valid():
                data = serializer.validated_data
                category_name = data.pop('category', '')
                category = None
                if category_name:
                    category, _ = Category.objects.get_or_create(
                        name=category_name,
                        defaults={'slug': category_name.lower().replace(' ', '-')}
                    )
                try:
                    Product.objects.create(
                        seller=seller,
                        category=category,
                        **data,
                    )
                    created.append(row_data['SKU'])
                except Exception as e:
                    errors.append({'row': row_num, 'SKU': row_data.get('SKU'), 'error': str(e)})
            else:
                errors.append({'row': row_num, 'data': row_data, 'errors': serializer.errors})

        return Response({
            'status': 'success',
            'created_count': len(created),
            'error_count': len(errors),
            'created_skus': created,
            'errors': errors,
        }, status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED)

